import os
import csv
import json
from typing import Protocol

from openpyxl import load_workbook


class BillProcessor(Protocol):
    def calculate_total(self) -> float:
        ...


class JsonBillProcessor:
    def __init__(self, filename):
        self.filename = filename
        self._doc = self._load_doc()

    def _load_doc(self):
        with open(self.filename, "r") as f:
            doc = json.load(f)
        return doc

    def calculate_total(self) -> float:
        return sum(row["Quantity"] * row["Price"] for row in self._doc)


class CsvBillProcessor:
    def __init__(self, filename):
        self.filename = filename
        self._rows = self._load_rows()

    def _load_rows(self):
        rows = []
        with open(self.filename, "r") as f:
            rd = csv.DictReader(f)
            for row in rd:
                rows.append(row)
        return rows

    def calculate_total(self) -> float:
        return sum(float(row["Qty"]) * float(row["Price"]) for row in self._doc)


class XlsxBillProcessor:
    def __init__(self, filename):
        self.filename = filename
        self._rows = self._load_workbook()

    def _load_workbook(self):
        workbook = load_workbook(self.filename)
        ws = workbook.worksheets[0]
        header = []
        rows = []
        first = True
        for row in ws.iter_rows():
            if first:
                for cell in row:
                    header.append(cell.value)
                first = False
                continue
            values = []
            for cell in row:
                values.append(cell.value)
            rows.append(values)

        results = []
        for row in rows:
            results.append(dict(zip(header, row)))
        return results

    def calculate_total(self) -> float:
        return sum(row["qty"] * row["Price"] for row in self._doc)


class BillProcessorFactory:
    def create_bill_processor(self, filename):
        _, ext = os.path.splitext(filename)
        ext = ext[1:].title()
        classname = f"{ext}BillProcessor"
        BillProcessorClass = globals().get(classname)
        if BillProcessorClass is None:
            raise Exception("Unsupported format")
        return BillProcessorClass(filename)


def calculate_total_monthly_bills(bills: list[BillProcessor]) -> float:
    return sum(bill.calculate_total() for bill in bills)
